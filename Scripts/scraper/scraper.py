import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support import expected_conditions as EC
import os
import csv
import time
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
load_dotenv()



class Scraper:
    def __init__(self, headless = False):
        self.options = uc.ChromeOptions()
        if headless:
            self.options.add_argument('--headless')  
        self.options.add_argument('--disable-gpu')
        self.options.add_argument('--no-sandbox')
        self.options.add_argument('--disable-dev-shm-usage')
        self.options.add_argument('--remote-debugging-port=9222')  
        self.driver = uc.Chrome(options=self.options)

    def open_page(self, url):
        """Open the page and wait for the content to load."""
        self.driver.get(url)
        time.sleep(5)
        print("Nicely Opened")

    def SaisirInfo(self):
        """Set Information Into Fields."""
        email_xpath = '//*[@id="email"]'
        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, email_xpath))
        )

        email = self.driver.find_element(By.XPATH, email_xpath)
        email.send_keys(os.getenv("GMAIL"))

    
        password_xpath = '//*[@id="password"]'    
        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, password_xpath))
        )
        password = self.driver.find_element(By.XPATH, password_xpath)
        password.send_keys(os.getenv("PASS"))

        time.sleep(2)

    def Clik_Button(self):
        BUTTON_PATH = '//*[@id="kt_login_form"]/div[4]/button'  
        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.XPATH, BUTTON_PATH))
        )
        BUTTON = self.driver.find_element(By.XPATH, BUTTON_PATH)
        BUTTON.click()
        time.sleep(2)


    def _toCountry(self, link):
        self.driver.get(link)
        time.sleep(2)


    def getListOfProducts_per_country(self, link):
        product_links = []
        self._toCountry(link)
        time.sleep(15)
    
        listing_xpath = '//*[@id="page-content-wrapper"]/div[2]/div/div[2]/div/div/div[2]/div[3]'
        WebDriverWait(self.driver, 70).until(
            EC.presence_of_element_located((By.XPATH, listing_xpath))
        )
    
        listing_container = self.driver.find_element(By.XPATH, listing_xpath)
    
        #product_class = 'col-12 col-sm-6 col-md-6 col-lg-6 col-xl-3'
    

        WebDriverWait(self.driver, 70).until(
            EC.presence_of_all_elements_located((By.CLASS_NAME, 'col-12')) 
        )
    

        product_elements = listing_container.find_elements(By.CLASS_NAME, 'col-12')

        for product in product_elements : 
            link = product.find_element(By.TAG_NAME, 'a')
            product_links.append(link.get_attribute("href"))

    
        return product_links
    
    def dirigerVersProduct(self, link):
        self.driver.get(link)


    def getDetailsOfProduct(self, link):
        self.dirigerVersProduct(link)

        WebDriverWait(self.driver, 30).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'row'))
        )

        try:
            name = self.driver.find_element(By.CLASS_NAME, 'fs-4.mb-0.text-uppercase.mt-2.mt-md-0.fw-bold').text
        except NoSuchElementException:
            name = "Nom introuvable"

        try:
            cate = self.driver.find_element(By.XPATH, '//*[@id="page-content-wrapper"]/div[2]/div/div[3]/div/div[1]/div[2]/div[2]/div[2]/div[1]/div[1]/a/small').text
        except NoSuchElementException:
            cate = "Catégorie introuvable"

        try:
            price_element = self.driver.find_element(By.CLASS_NAME, 'title.text-dark.me-2')
            price = price_element.text
        except NoSuchElementException:
            price = "Prix introuvable"

        try:
            facts = self.driver.find_elements(By.CLASS_NAME, 'text-primary.fw-bold.fa-lg.mb-0')
            price_fact = facts[0].text if len(facts) > 0 else "N/A"
            quant = facts[1].text if len(facts) > 1 else "N/A"
        except NoSuchElementException:
            price_fact = "N/A"
            quant = "N/A"

        return name, cate, price, price_fact, quant, link
    
    def save_product_data(self, data, csv_file="produits.csv", xlsx_file="produits.xlsx"):
        import csv
        from openpyxl import Workbook, load_workbook

        headers = ["Product Name", "Category", "Displayed Price", "Factured Price", "Quantity", "Product Link"]


        # --- CSV ---
        file_exists = os.path.isfile(csv_file)
        with open(csv_file, mode="a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if not file_exists:
                writer.writerow(headers)
            writer.writerow(data)

        # --- Excel ---
        if os.path.exists(xlsx_file):
            wb = load_workbook(xlsx_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(headers)

        ws.append(data)
        wb.save(xlsx_file)




